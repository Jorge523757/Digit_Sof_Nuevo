"""
DIGT SOFT - Vistas del Módulo de Compras
"""

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db.models import Q, Sum, Count
from datetime import datetime
from .models import Compra, DetalleCompra
from .forms import CompraForm
from usuarios.decorators import staff_required
from .forms import CompraForm, BuscarCompraForm
from usuarios.decorators import staff_required


@login_required
@staff_required
def compras_lista(request):
    """Lista de compras con filtros avanzados"""
    compras = Compra.objects.select_related('proveedor', 'usuario').all().order_by('-fecha_compra')

    # Búsqueda general
    busqueda = request.GET.get('busqueda', '')
    if busqueda:
        compras = compras.filter(
            Q(numero_compra__icontains=busqueda) |
            Q(proveedor__nombre_empresa__icontains=busqueda) |
            Q(proveedor__ruc__icontains=busqueda)
        )

    # Filtro por estado
    estado = request.GET.get('estado', '')
    if estado:
        compras = compras.filter(estado=estado)

    # Filtro por proveedor
    proveedor_id = request.GET.get('proveedor', '')
    if proveedor_id:
        compras = compras.filter(proveedor_id=proveedor_id)

    # Filtro por método de pago
    metodo_pago = request.GET.get('metodo_pago', '')
    if metodo_pago:
        compras = compras.filter(metodo_pago=metodo_pago)

    # Filtro por usuario
    usuario_id = request.GET.get('usuario', '')
    if usuario_id:
        compras = compras.filter(usuario_id=usuario_id)

    # Filtro por rango de fechas
    fecha_desde = request.GET.get('fecha_desde', '')
    fecha_hasta = request.GET.get('fecha_hasta', '')

    if fecha_desde:
        try:
            fecha_desde_obj = datetime.strptime(fecha_desde, '%Y-%m-%d')
            compras = compras.filter(fecha_compra__date__gte=fecha_desde_obj.date())
        except ValueError:
            pass

    if fecha_hasta:
        try:
            fecha_hasta_obj = datetime.strptime(fecha_hasta, '%Y-%m-%d')
            compras = compras.filter(fecha_compra__date__lte=fecha_hasta_obj.date())
        except ValueError:
            pass

    # Paginación
    paginator = Paginator(compras, 15)
    page_obj = paginator.get_page(request.GET.get('page'))

    # Estadísticas generales
    total_compras = Compra.objects.count()
    completadas = Compra.objects.filter(estado='COMPLETADA').count()
    pendientes = Compra.objects.filter(estado='PENDIENTE').count()
    aprobadas = Compra.objects.filter(estado='APROBADA').count()
    total_gastado = Compra.objects.filter(estado='COMPLETADA').aggregate(
        total=Sum('total')
    )['total'] or 0

    # Estadísticas del filtro actual
    compras_filtradas = compras.aggregate(
        total=Sum('total'),
        count=Count('id')
    )

    # Obtener listas para los filtros
    from proveedores.models import Proveedor
    from django.contrib.auth.models import User

    proveedores = Proveedor.objects.filter(activo=True).order_by('nombre_empresa')
    usuarios = User.objects.filter(compras_realizadas__isnull=False).distinct()
    estados = Compra.ESTADO_CHOICES
    metodos_pago = Compra.METODO_PAGO_CHOICES

    context = {
        'page_obj': page_obj,
        'total_compras': total_compras,
        'completadas': completadas,
        'pendientes': pendientes,
        'aprobadas': aprobadas,
        'total_gastado': total_gastado,
        'compras_filtradas_count': compras_filtradas['count'] or 0,
        'compras_filtradas_total': compras_filtradas['total'] or 0,
        'busqueda': busqueda,
        'estado': estado,
        'proveedor_id': proveedor_id,
        'metodo_pago': metodo_pago,
        'usuario_id': usuario_id,
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
        'proveedores': proveedores,
        'usuarios': usuarios,
        'estados': estados,
        'metodos_pago': metodos_pago,
    }
    return render(request, 'compras/lista.html', context)


@login_required
@staff_required
def compra_crear(request):
    """Crear nueva compra"""
    if request.method == 'POST':
        form = CompraForm(request.POST)
        if form.is_valid():
            compra = form.save(commit=False)
            compra.usuario = request.user
            compra.save()
            messages.success(request, f'✅ Compra "{compra.numero_compra}" creada exitosamente.')
            return redirect('compras:detalle', pk=compra.pk)
    else:
        form = CompraForm()

    return render(request, 'compras/form.html', {
        'form': form,
        'titulo': 'Registrar Nueva Compra',
        'accion': 'Crear'
    })


@login_required
@staff_required
def compra_detalle(request, pk):
    """Ver detalle de compra"""
    compra = get_object_or_404(Compra, pk=pk)
    detalles = compra.items.select_related('producto').all()

    return render(request, 'compras/detalle.html', {
        'compra': compra,
        'detalles': detalles,
    })


@login_required
@staff_required
def compra_editar(request, pk):
    """Editar compra"""
    compra = get_object_or_404(Compra, pk=pk)

    if request.method == 'POST':
        form = CompraForm(request.POST, instance=compra)
        if form.is_valid():
            compra = form.save()
            messages.success(request, f'✅ Compra actualizada correctamente.')
            return redirect('compras:detalle', pk=compra.pk)
    else:
        form = CompraForm(instance=compra)

    return render(request, 'compras/form.html', {
        'form': form,
        'compra': compra,
        'titulo': 'Editar Compra',
        'accion': 'Actualizar'
    })


# REPORTES PDF Y EXCEL

    if request.method == 'POST':
        form = CompraForm(request.POST, instance=compra)
        if form.is_valid():
            compra = form.save()
            messages.success(request, f'✅ Compra actualizada correctamente.')
            return redirect('compras:detalle', pk=compra.pk)
    else:
        form = CompraForm(instance=compra)

    return render(request, 'compras/form.html', {
        'form': form,
        'compra': compra,
        'titulo': 'Editar Compra',
        'accion': 'Actualizar'
    })


# REPORTES PDF Y EXCEL
@login_required
@staff_required
def compra_reporte_pdf(request):
    """Generar reporte de compras en PDF"""
    from django.http import HttpResponse
    from django.template.loader import get_template
    from xhtml2pdf import pisa
    from io import BytesIO

    # Obtener filtros
    estado = request.GET.get('estado', '')
    fecha_desde = request.GET.get('fecha_desde', '')
    fecha_hasta = request.GET.get('fecha_hasta', '')

    compras = Compra.objects.select_related('proveedor', 'usuario').all()

    if estado:
        compras = compras.filter(estado=estado)
    if fecha_desde:
        try:
            fecha_desde_obj = datetime.strptime(fecha_desde, '%Y-%m-%d')
            compras = compras.filter(fecha_compra__date__gte=fecha_desde_obj.date())
        except ValueError:
            pass
    if fecha_hasta:
        try:
            fecha_hasta_obj = datetime.strptime(fecha_hasta, '%Y-%m-%d')
            compras = compras.filter(fecha_compra__date__lte=fecha_hasta_obj.date())
        except ValueError:
            pass

    compras = compras.order_by('-fecha_compra')

    context = {
        'compras': compras,
        'fecha': datetime.now(),
        'usuario': request.user,
        'total_compras': compras.count(),
        'total_gastado': compras.aggregate(Sum('total'))['total__sum'] or 0,
    }

    template = get_template('compras/reporte_pdf.html')
    html = template.render(context)

    result = BytesIO()
    pdf = pisa.pisaDocument(BytesIO(html.encode("UTF-8")), result)

    if not pdf.err:
        response = HttpResponse(result.getvalue(), content_type='application/pdf')
        filename = f'reporte_compras_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    messages.error(request, 'Error al generar el PDF')
    return redirect('compras:lista')


@login_required
@staff_required
def compra_reporte_excel(request):
    """Generar reporte de compras en Excel"""
    from django.http import HttpResponse
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill

    # Obtener filtros
    estado = request.GET.get('estado', '')
    fecha_desde = request.GET.get('fecha_desde', '')
    fecha_hasta = request.GET.get('fecha_hasta', '')

    compras = Compra.objects.select_related('proveedor', 'usuario').all()

    if estado:
        compras = compras.filter(estado=estado)
    if fecha_desde:
        try:
            fecha_desde_obj = datetime.strptime(fecha_desde, '%Y-%m-%d')
            compras = compras.filter(fecha_compra__date__gte=fecha_desde_obj.date())
        except ValueError:
            pass
    if fecha_hasta:
        try:
            fecha_hasta_obj = datetime.strptime(fecha_hasta, '%Y-%m-%d')
            compras = compras.filter(fecha_compra__date__lte=fecha_hasta_obj.date())
        except ValueError:
            pass

    compras = compras.order_by('-fecha_compra')

    # Crear workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte de Compras"

    # Estilos
    header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    # Encabezados
    headers = ['Nº Compra', 'Proveedor', 'Usuario', 'Fecha', 'Total', 'Estado', 'Método Pago']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # Datos
    for row_num, compra in enumerate(compras, 2):
        ws.cell(row=row_num, column=1, value=compra.numero_compra)
        ws.cell(row=row_num, column=2, value=compra.proveedor.nombre_empresa)
        ws.cell(row=row_num, column=3, value=compra.usuario.get_full_name() if compra.usuario else 'N/A')
        ws.cell(row=row_num, column=4, value=compra.fecha_compra.strftime('%d/%m/%Y'))
        ws.cell(row=row_num, column=5, value=float(compra.total))
        ws.cell(row=row_num, column=6, value=compra.get_estado_display())
        ws.cell(row=row_num, column=7, value=compra.get_metodo_pago_display())

    # Ajustar anchos
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Respuesta
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f'reporte_compras_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)

    return response

    if request.method == 'POST':
        numero = compra.numero_compra
        compra.delete()
        messages.success(request, f'✅ Compra "{numero}" eliminada correctamente.')
        return redirect('compras:lista')

    return render(request, 'compras/eliminar.html', {'compra': compra})


# REPORTES PDF Y EXCEL
# ==============================================

from django.contrib.auth.decorators import login_required
from usuarios.decorators import staff_required

@login_required
@staff_required
def compra_reporte_pdf(request):
    """Generar reporte de compras en PDF"""
    from utils.reportes import generar_pdf
    from datetime import datetime

    query = request.GET.get('q', '').strip()
    estado = request.GET.get('estado', '')

    datos = Compra.objects.all().select_related('proveedor')

    if query:
        datos = datos.filter(
            Q(numero_compra__icontains=query) |
            Q(proveedor__nombre_empresa__icontains=query)
        )

    if estado:
        datos = datos.filter(estado=estado)

    datos = datos.order_by('-fecha_compra')

    context = {
        'datos': datos,
        'fecha': datetime.now(),
        'usuario': request.user,
        'total': datos.count(),
        'total_gastado': datos.aggregate(Sum('total'))['total__sum'] or 0,
    }

    filename = f'reporte_compras_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
    return generar_pdf('reportes/compras_pdf.html', context, filename)


@login_required
@staff_required
def compra_reporte_excel(request):
    """Generar reporte de compras en Excel"""
    from utils.reportes import generar_excel_avanzado
    from datetime import datetime

    query = request.GET.get('q', '').strip()
    estado = request.GET.get('estado', '')

    datos_query = Compra.objects.all().select_related('proveedor')

    if query:
        datos_query = datos_query.filter(
            Q(numero_compra__icontains=query) |
            Q(proveedor__nombre_empresa__icontains=query)
        )

    if estado:
        datos_query = datos_query.filter(estado=estado)

    datos_query = datos_query.order_by('-fecha_compra')

    datos = []
    for item in datos_query:
        datos.append({
            'numero_compra': item.numero_compra,
            'proveedor': str(item.proveedor) if item.proveedor else '',
            'fecha_compra': item.fecha_compra,
            'subtotal': float(item.subtotal),
            'impuestos': float(item.impuestos),
            'total': float(item.total),
            'estado': item.get_estado_display(),
        })

    columnas = [
        ('numero_compra', 'Número Compra', 'texto'),
        ('proveedor', 'Proveedor', 'texto'),
        ('fecha_compra', 'Fecha', 'fecha'),
        ('subtotal', 'Subtotal', 'moneda'),
        ('impuestos', 'Impuestos', 'moneda'),
        ('total', 'Total', 'moneda'),
        ('estado', 'Estado', 'texto'),
    ]

    titulo = 'Reporte de Compras'
    filename = f'reporte_compras_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

    totales = ['subtotal', 'impuestos', 'total']

    return generar_excel_avanzado(datos, columnas, titulo, filename, totales=totales)
