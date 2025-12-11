"""
Utilidades para generar reportes en PDF y Excel
"""
from django.http import HttpResponse
from django.template.loader import get_template
from xhtml2pdf import pisa
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from io import BytesIO


def generar_pdf(template_name, context, filename="reporte.pdf"):
    """
    Genera un archivo PDF a partir de un template HTML

    Args:
        template_name: Nombre del template HTML
        context: Contexto con datos para el template
        filename: Nombre del archivo PDF a generar

    Returns:
        HttpResponse con el PDF generado
    """
    template = get_template(template_name)
    html = template.render(context)

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    # Crear PDF
    pisa_status = pisa.CreatePDF(html, dest=response)

    if pisa_status.err:
        return HttpResponse('Error al generar el PDF', status=500)

    return response


def generar_excel(datos, columnas, titulo, filename="reporte.xlsx"):
    """
    Genera un archivo Excel con formato profesional

    Args:
        datos: Lista de diccionarios con los datos
        columnas: Lista de tuplas (clave, nombre_columna)
        titulo: Título del reporte
        filename: Nombre del archivo Excel a generar

    Returns:
        HttpResponse con el Excel generado
    """
    # Crear workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = titulo[:31]  # Excel limita a 31 caracteres

    # Estilos
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")

    title_fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
    title_font = Font(bold=True, color="FFFFFF", size=16)
    title_alignment = Alignment(horizontal="center", vertical="center")

    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Título principal
    ws.merge_cells(f'A1:{get_column_letter(len(columnas))}1')
    title_cell = ws['A1']
    title_cell.value = titulo
    title_cell.fill = title_fill
    title_cell.font = title_font
    title_cell.alignment = title_alignment
    ws.row_dimensions[1].height = 30

    # Información adicional
    ws.merge_cells(f'A2:{get_column_letter(len(columnas))}2')
    info_cell = ws['A2']
    info_cell.value = f"Generado el: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
    info_cell.alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 20

    # Encabezados
    row_num = 4
    for col_num, (_, col_name) in enumerate(columnas, 1):
        cell = ws.cell(row=row_num, column=col_num)
        cell.value = col_name
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border

    # Datos
    for row_data in datos:
        row_num += 1
        for col_num, (key, _) in enumerate(columnas, 1):
            cell = ws.cell(row=row_num, column=col_num)
            value = row_data.get(key, '')

            # Formatear valores especiales
            if isinstance(value, datetime):
                value = value.strftime('%d/%m/%Y %H:%M:%S')
            elif isinstance(value, bool):
                value = 'Sí' if value else 'No'
            elif value is None:
                value = ''

            cell.value = value
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="center")

    # Ajustar ancho de columnas
    for col_num in range(1, len(columnas) + 1):
        column_letter = get_column_letter(col_num)
        max_length = 0
        for row in ws[column_letter]:
            try:
                if len(str(row.value)) > max_length:
                    max_length = len(str(row.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Preparar respuesta
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    return response


def generar_excel_avanzado(datos, columnas, titulo, filename="reporte.xlsx",
                          totales=None, graficos=None):
    """
    Genera un archivo Excel con características avanzadas (totales, gráficos)

    Args:
        datos: Lista de diccionarios con los datos
        columnas: Lista de tuplas (clave, nombre_columna, tipo)
        titulo: Título del reporte
        filename: Nombre del archivo
        totales: Dict con columnas que deben tener totales
        graficos: Lista de configuraciones de gráficos

    Returns:
        HttpResponse con el Excel generado
    """
    # Crear workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = titulo[:31]

    # Estilos
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    total_fill = PatternFill(start_color="E8F8F5", end_color="E8F8F5", fill_type="solid")
    total_font = Font(bold=True, size=11)

    # Título
    ws.merge_cells(f'A1:{get_column_letter(len(columnas))}1')
    title_cell = ws['A1']
    title_cell.value = titulo
    title_cell.fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
    title_cell.font = Font(bold=True, color="FFFFFF", size=16)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # Fecha
    ws.merge_cells(f'A2:{get_column_letter(len(columnas))}2')
    ws['A2'].value = f"Generado el: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
    ws['A2'].alignment = Alignment(horizontal="center")

    # Encabezados
    row_num = 4
    for col_num, (_, col_name, *_) in enumerate(columnas, 1):
        cell = ws.cell(row=row_num, column=col_num)
        cell.value = col_name
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Datos
    for row_data in datos:
        row_num += 1
        for col_num, col_info in enumerate(columnas, 1):
            key = col_info[0]
            cell = ws.cell(row=row_num, column=col_num)
            value = row_data.get(key, '')

            # Formatear según tipo
            if len(col_info) > 2:
                tipo = col_info[2]
                if tipo == 'moneda' and value:
                    cell.value = float(value) if value else 0
                    cell.number_format = '$#,##0.00'
                elif tipo == 'numero' and value:
                    cell.value = float(value) if value else 0
                    cell.number_format = '#,##0'
                elif tipo == 'fecha' and value:
                    if isinstance(value, str):
                        cell.value = value
                    else:
                        cell.value = value.strftime('%d/%m/%Y')
                else:
                    cell.value = value
            else:
                cell.value = value

    # Totales
    if totales:
        row_num += 1
        for col_num, col_info in enumerate(columnas, 1):
            cell = ws.cell(row=row_num, column=col_num)
            key = col_info[0]

            if key in totales:
                if col_num == 1:
                    cell.value = "TOTAL:"
                    cell.font = total_font
                else:
                    # Calcular suma
                    start_row = 5
                    end_row = row_num - 1
                    cell.value = f"=SUM({get_column_letter(col_num)}{start_row}:{get_column_letter(col_num)}{end_row})"
                    if len(col_info) > 2 and col_info[2] == 'moneda':
                        cell.number_format = '$#,##0.00'
                    cell.font = total_font

            cell.fill = total_fill

    # Ajustar anchos
    for col_num in range(1, len(columnas) + 1):
        ws.column_dimensions[get_column_letter(col_num)].width = 15

    # Guardar
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    return response

